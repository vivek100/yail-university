"""Build the compact AFC audit population used by acct-audit-easy.

The task stays grounded in the authentic GDPval source workbook but trims the
population so small open models can read it inside a 64K context window.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


HERE = Path(__file__).resolve().parent
SOURCE = HERE.parent / "acct-afc-audit-sampling" / "reference_files" / "Population v2.xlsx"
OUT_DIR = HERE / "reference_files"
OUT_WORKBOOK = OUT_DIR / "Population easy.xlsx"
OUT_PROVENANCE = OUT_DIR / "provenance.json"

HIGH_RISK_COUNTRIES = {"Cayman Islands", "Pakistan", "UAE"}
VARIANCE_THRESHOLD = 0.15
TARGET_ROWS = 120


def _to_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _risk_flags(row: dict[str, Any]) -> dict[str, bool]:
    q3 = _to_float(row.get("Q3 2024 KRI"))
    q2 = _to_float(row.get("Q2 2024 KRI"))
    variance_flag = False
    if q2 not in (None, 0) and q3 is not None:
        variance_flag = abs((q3 - q2) / q2) > VARIANCE_THRESHOLD
    return {
        "high_risk_country": str(row.get("Country", "")).strip() in HIGH_RISK_COUNTRIES,
        "zero_value": q3 == 0 or q2 == 0,
        "variance": variance_flag,
    }


def _risk_score(row: dict[str, Any]) -> tuple[int, int, int]:
    flags = _risk_flags(row)
    score = 0
    score += 4 if flags["high_risk_country"] else 0
    score += 3 if flags["variance"] else 0
    score += 2 if flags["zero_value"] else 0
    no = int(float(row["No"]))
    return (-score, no, len(str(row.get("KRIs", ""))))


def _load_source() -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    header = [str(cell).strip() for cell in rows[0]]
    records = [dict(zip(header, row)) for row in rows[1:]]
    return header, records


def _select_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected: dict[int, dict[str, Any]] = {}

    def add(row: dict[str, Any]) -> None:
        selected[int(float(row["No"]))] = row

    # Guarantee coverage before filling by risk score.
    for division in sorted({str(row.get("Division", "")).strip() for row in records}):
        candidates = [row for row in records if str(row.get("Division", "")).strip() == division]
        for row in sorted(candidates, key=_risk_score)[:8]:
            add(row)

    for country in sorted(HIGH_RISK_COUNTRIES):
        candidates = [row for row in records if str(row.get("Country", "")).strip() == country]
        for row in sorted(candidates, key=_risk_score)[:10]:
            add(row)

    for row in sorted(records, key=_risk_score):
        add(row)
        if len(selected) >= TARGET_ROWS:
            break

    return [selected[key] for key in sorted(selected)]


def _write_workbook(header: list[str], rows: list[dict[str, Any]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Population"
    ws.append(header)
    for row in rows:
        ws.append([row.get(column) for column in header])
    for cell in ws[1]:
        cell.style = "Headline 4"
    wb.save(OUT_WORKBOOK)
    wb.close()


def main() -> int:
    header, records = _load_source()
    selected = _select_records(records)
    _write_workbook(header, selected)
    flags = [_risk_flags(row) for row in selected]
    provenance = {
        "source": str(SOURCE.relative_to(HERE.parent.parent)),
        "source_rows": len(records),
        "output": OUT_WORKBOOK.name,
        "output_rows": len(selected),
        "variance_threshold": VARIANCE_THRESHOLD,
        "high_risk_countries": sorted(HIGH_RISK_COUNTRIES),
        "risk_flag_counts": {
            "high_risk_country": sum(flag["high_risk_country"] for flag in flags),
            "zero_value": sum(flag["zero_value"] for flag in flags),
            "variance": sum(flag["variance"] for flag in flags),
        },
        "division_count": len({row.get("Division") for row in selected}),
        "country_count": len({row.get("Country") for row in selected}),
        "row_ids": [int(float(row["No"])) for row in selected],
    }
    OUT_PROVENANCE.write_text(json.dumps(provenance, indent=2), encoding="utf-8")
    print(json.dumps(provenance, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
