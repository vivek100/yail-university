"""HUD environment for GDPval-style knowledge-work tasks.

Frozen, CPU-only, offline. The harness stages a bundle of reference materials
into the solver workspace, hands the agent a natural-language brief, and on
completion grades a native professional deliverable (.xlsx / .docx / .pptx /
.pdf / code) with a plain, readable grader.

Grading is fully transparent: there is no sealing or encryption. On completion
the harness loads grader/rubric from plaintext task args produced by `tasks/`
and calls `grade(workspace, deliverable, rubric)`. The author-local `_hidden/`
directory is deliberately excluded from the deployed image and from the solver
workspace; it is a source file for task sync, not runtime state.
"""

from __future__ import annotations

import asyncio
import contextlib
import inspect
import importlib
import importlib.util
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import tomllib
import types
from pathlib import Path
from typing import Any, AsyncGenerator, Literal

EditCommand = Literal["view", "create", "str_replace", "insert", "undo_edit"]

APP_ROOT = Path(__file__).resolve().parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))  # so task graders can import local helpers

try:
    from fastmcp import FastMCP
    from hud import Environment
    from hud.capabilities import Capability
    from hud.graders import EvaluationResult, SubScore

    HUD_AVAILABLE = True
except (ImportError, ModuleNotFoundError):  # pragma: no cover - lets env.py import bare
    HUD_AVAILABLE = False
    EvaluationResult = None  # type: ignore[assignment]
    SubScore = None  # type: ignore[assignment]

    class ToolError(RuntimeError):
        pass

    class Environment:  # type: ignore[no-redef]
        def __init__(self, name: str) -> None:
            self.name = name

        def template(self, _func=None, **_kwargs):
            def decorator(func):
                return func

            return decorator if _func is None else decorator(_func)

        def initialize(self, func):
            return func

        def shutdown(self, func):
            return func
else:

    class ToolError(RuntimeError):
        pass


WORKSPACE_DIR = Path(
    os.environ.get("WORKSPACE_DIR", "/workspace/target" if Path("/workspace").exists() else "/tmp/gdpval_workspace")
).resolve()
WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)
WORKSPACE_BASH_STREAM_LIMIT = 32 * 1024 * 1024

# Stripped from the solver shell so the agent cannot reach the judge or any key.
SOLVER_ENV_DROP_NAMES = {
    "HUD_API_KEY", "HUD_API_URL", "HUD_GATEWAY_URL",
    "GDPVAL_JUDGE_API_KEY", "GDPVAL_JUDGE_BASE_URL",
    "OPENAI_API_KEY", "OPENAI_BASE_URL", "ANTHROPIC_API_KEY", "GEMINI_API_KEY", "GOOGLE_API_KEY",
}
SOLVER_ENV_SECRET_MARKERS = ("API_KEY", "TOKEN", "SECRET", "PASSWORD", "CREDENTIAL")
# Blocked in solver shell commands to keep the evaluation offline.
EGRESS_PATTERNS = (
    r"\b(?:curl|wget|ssh|scp|sftp|ftp|nc|ncat|telnet|rsync)\b",
    r"\b(?:pip|pip3)\s+install\b",
    r"\bpython(?:3(?:\.\d+)?)?\s+-m\s+pip\s+install\b",
    r"\buv\s+(?:pip\s+)?(?:add|install|sync)\b",
    r"\bgit\s+(?:clone|fetch|pull|ls-remote|submodule\s+update)\b",
    r"\b(?:http|https|ftp|ssh)://",
    r"/dev/tcp/",
)


def _load_env_name() -> str:
    override = os.environ.get("GDPVAL_HUD_ENV", "").strip()
    if override:
        return override
    config_path = APP_ROOT / "config.toml"
    if config_path.is_file():
        with config_path.open("rb") as handle:
            hud = tomllib.load(handle).get("hud", {})
        name = hud.get("production_environment") or hud.get("environment")
        if isinstance(name, str) and name.strip():
            return name.strip()
    return "gdpval-template"


ENV_NAME = _load_env_name()
env = Environment(name=ENV_NAME)
env.workspace(str(WORKSPACE_DIR))

tool_server = FastMCP(name="gdpval-spreadsheet-tools") if HUD_AVAILABLE else None
_MCP_SERVER_TASK: asyncio.Task[None] | None = None
_MCP_PORT: int | None = None


def _reject_solver_egress(command: str) -> None:
    lowered = command.lower()
    hits = [pat for pat in EGRESS_PATTERNS if re.search(pat, lowered)]
    if hits:
        raise ToolError(
            f"Network/egress is blocked in this offline evaluation (matched: {', '.join(hits)}). "
            "Work from the staged workspace artifacts."
        )


def _solver_subprocess_env() -> dict[str, str]:
    clean = dict(os.environ)
    for name in list(clean):
        if name in SOLVER_ENV_DROP_NAMES or any(marker in name for marker in SOLVER_ENV_SECRET_MARKERS):
            clean.pop(name, None)
    clean.update(
        {
            "HOME": "/home/solver" if Path("/home/solver").exists() else str(WORKSPACE_DIR),
            "USER": "solver",
            "LOGNAME": "solver",
            "XDG_CACHE_HOME": str(WORKSPACE_DIR / ".cache"),
            "TMPDIR": "/tmp",
        }
    )
    return clean


def _chown_solver_workspace() -> None:
    if os.getuid() != 0:
        return
    try:
        subprocess.run(["chown", "-R", "1000:1000", str(WORKSPACE_DIR)], check=True, capture_output=True)
    except (FileNotFoundError, subprocess.CalledProcessError) as exc:  # pragma: no cover
        print(f"[gdpval-env] warning: chown failed: {exc}", file=sys.stderr)


def _deliverable_path(path_text: str) -> Path:
    rel = Path(path_text or "deliverable/report.md")
    if rel.is_absolute() or ".." in rel.parts:
        raise RuntimeError(f"unsafe deliverable path: {path_text!r}")
    return WORKSPACE_DIR / rel


def _workspace_path(path_text: str) -> Path:
    rel = Path(path_text)
    if rel.is_absolute():
        candidate = rel.resolve()
    else:
        candidate = (WORKSPACE_DIR / rel).resolve()
    workspace = WORKSPACE_DIR.resolve()
    if candidate != workspace and workspace not in candidate.parents:
        raise ToolError(f"path must stay inside the workspace: {path_text!r}")
    return candidate


def _json_safe_cell(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _xlsx_records(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    return [dict(zip(header, row)) for row in rows[1:]]


def _as_number(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _row_id(value: Any) -> int | None:
    numeric = _as_number(value)
    if numeric is None:
        return None
    return int(numeric)


def _variance_fields(record: dict[str, Any]) -> tuple[float | None, float | None]:
    q3 = _as_number(record.get("Q3 2024 KRI"))
    q2 = _as_number(record.get("Q2 2024 KRI"))
    absolute = None if q3 is None or q2 is None else q3 - q2
    percent = None if absolute is None or q2 in (None, 0) else absolute / q2
    return absolute, percent


def _free_port() -> int:
    sock = socket.socket()
    sock.bind(("", 0))
    port = int(sock.getsockname()[1])
    sock.close()
    return port


async def _wait_for_port(port: int, timeout: float = 15.0) -> None:
    loop = asyncio.get_running_loop()
    deadline = loop.time() + timeout
    while True:
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=0.2):
                return
        except OSError:
            if loop.time() > deadline:
                raise TimeoutError(f"MCP server did not start on port {port}")
            await asyncio.sleep(0.1)


if tool_server is not None:

    @tool_server.tool
    async def read_xlsx(path: str, max_rows_per_sheet: int = 200) -> str:
        """Read a workspace .xlsx file as JSON rows.

        Args:
            path: Workspace-relative path, e.g. reference_files/Population easy.xlsx.
            max_rows_per_sheet: Maximum rows to return from each sheet, including header.
        """
        workbook_path = _workspace_path(path)
        if workbook_path.suffix.lower() != ".xlsx":
            raise ToolError("read_xlsx only accepts .xlsx files")
        if not workbook_path.is_file():
            raise ToolError(f"file not found: {path}")

        from openpyxl import load_workbook

        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        try:
            sheets: dict[str, dict[str, Any]] = {}
            limit = max(1, min(int(max_rows_per_sheet), 500))
            for ws in wb.worksheets:
                rows = []
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx >= limit:
                        break
                    rows.append([_json_safe_cell(cell) for cell in row])
                sheets[ws.title] = {
                    "returned_rows": len(rows),
                    "total_rows": ws.max_row,
                    "total_columns": ws.max_column,
                    "rows": rows,
                }
        finally:
            wb.close()
        return json.dumps({"path": path, "sheets": sheets}, ensure_ascii=True)

    @tool_server.tool
    async def write_xlsx(path: str, sheets: dict[str, list[list[Any]]]) -> str:
        """Write a workspace .xlsx file from JSON rows.

        Args:
            path: Workspace-relative output path, e.g. deliverable/Sample.xlsx.
            sheets: Mapping of sheet name to 2D rows. The first row should be headers.
        """
        output_path = _workspace_path(path)
        if output_path.suffix.lower() != ".xlsx":
            raise ToolError("write_xlsx only writes .xlsx files")
        if not isinstance(sheets, dict) or not sheets:
            raise ToolError("sheets must be a non-empty object")

        from openpyxl import Workbook

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for raw_name, rows in sheets.items():
            name = str(raw_name).strip()[:31] or "Sheet"
            ws = wb.create_sheet(title=name)
            if not isinstance(rows, list):
                raise ToolError(f"sheet {name!r} must be a list of rows")
            for row in rows[:1000]:
                if not isinstance(row, list):
                    raise ToolError(f"sheet {name!r} contains a non-list row")
                ws.append([_json_safe_cell(cell) for cell in row[:100]])
        wb.save(output_path)
        wb.close()
        return json.dumps({"wrote": _display(output_path), "sheet_count": len(sheets)}, ensure_ascii=True)

    @tool_server.tool
    async def write_audit_sample_workbook(
        selected_ids: list[int],
        sample_size_note: str = "90% confidence, 10% tolerable error; selected a practical risk-based sample.",
        source_path: str = "reference_files/Population easy.xlsx",
        output_path: str = "deliverable/Sample.xlsx",
    ) -> str:
        """Create an AFC audit sample workbook from selected source row IDs.

        The tool materializes spreadsheet tabs; it does not choose the sample.
        Select 12-30 row IDs first, then call this tool with those IDs.
        """
        source = _workspace_path(source_path)
        output = _workspace_path(output_path)
        if source.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
            raise ToolError("source_path and output_path must be .xlsx files")
        records = _xlsx_records(source)
        if not records:
            raise ToolError(f"no records found in {source_path}")
        selected_set = {int(item) for item in selected_ids}
        if not selected_set:
            raise ToolError("selected_ids must not be empty")

        by_id = {row_id: record for record in records if (row_id := _row_id(record.get("No"))) is not None}
        missing = sorted(selected_set - set(by_id))
        if missing:
            raise ToolError(f"selected_ids not found in source workbook: {missing[:20]}")

        analysis_header = [
            "No",
            "Division",
            "Sub-Division",
            "Country",
            "Legal Entity",
            "KRIs",
            "Q3 2024 KRI",
            "Q2 2024 KRI",
            "QoQ Variance",
            "QoQ Variance %",
            "Selected",
        ]
        analysis_rows = [analysis_header]
        sample_rows = [analysis_header + ["Rationale"]]
        for record in records:
            row_id = _row_id(record.get("No"))
            absolute, percent = _variance_fields(record)
            selected = bool(row_id in selected_set)
            base = [
                row_id,
                record.get("Division"),
                record.get("Sub-Division"),
                record.get("Country"),
                record.get("Legal Entity"),
                record.get("KRIs"),
                record.get("Q3 2024 KRI"),
                record.get("Q2 2024 KRI"),
                absolute,
                percent,
                "Y" if selected else "",
            ]
            analysis_rows.append(base)
            if selected:
                flags = []
                q3 = _as_number(record.get("Q3 2024 KRI"))
                q2 = _as_number(record.get("Q2 2024 KRI"))
                if q3 == 0 or q2 == 0:
                    flags.append("zero Q2/Q3 value")
                if percent is not None and abs(percent) > 0.15:
                    flags.append("large QoQ variance")
                country = str(record.get("Country", "")).strip()
                if country in {"Cayman Islands", "Pakistan", "UAE"}:
                    flags.append(f"high-risk jurisdiction: {country}")
                rationale = "; ".join(flags) if flags else "coverage row selected for audit judgment"
                sample_rows.append(base + [rationale])

        sample_calc_rows = [
            ["Item", "Value"],
            ["Population rows", len(records)],
            ["Confidence level", "90%"],
            ["Tolerable error rate", "10%"],
            ["Method", sample_size_note],
            ["Selected sample size", len(selected_set)],
            ["Expected sample range", "12-30 rows"],
        ]

        from openpyxl import Workbook

        output.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for sheet_name, rows in {
            "Sample Size Calculation": sample_calc_rows,
            "Risk Analysis": analysis_rows,
            "Selected Sample": sample_rows,
        }.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append([_json_safe_cell(cell) for cell in row])
        wb.save(output)
        wb.close()
        return json.dumps(
            {
                "wrote": _display(output),
                "selected_count": len(selected_set),
                "selected_ids": sorted(selected_set),
                "sheets": ["Sample Size Calculation", "Risk Analysis", "Selected Sample"],
            },
            ensure_ascii=True,
        )


@env.initialize
async def _start_mcp_tools() -> None:
    global _MCP_PORT, _MCP_SERVER_TASK
    if tool_server is None:
        return
    if _MCP_SERVER_TASK is None:
        _MCP_PORT = _free_port()
        _MCP_SERVER_TASK = asyncio.create_task(
            tool_server.run_async(transport="http", host="127.0.0.1", port=_MCP_PORT, show_banner=False)
        )
        await _wait_for_port(_MCP_PORT)
    env.add_capability(Capability.mcp(name="gdpval-tools", url=f"http://127.0.0.1:{_MCP_PORT}/mcp"))


@env.shutdown
async def _stop_mcp_tools() -> None:
    global _MCP_SERVER_TASK
    if _MCP_SERVER_TASK is not None:
        _MCP_SERVER_TASK.cancel()
        with contextlib.suppress(asyncio.CancelledError):
            await _MCP_SERVER_TASK
        _MCP_SERVER_TASK = None


# Scratch dirs a confused agent may use instead of the workspace. These are listed
# only in diagnostics (never graded) so a human can see *where* a misplaced
# deliverable went without risking a false positive from a stale/cross-run file.
SCRATCH_ROOTS = ("/tmp", "/home/solver", "/root")
_WALK_SKIP_DIRS = {".git", "__pycache__", ".cache", ".config", ".ipython",
                   "node_modules", ".npm", "reference_files"}


def _within(root: Path, path: Path) -> bool:
    try:
        rp, rr = path.resolve(), root.resolve()
    except OSError:
        return False
    return rp == rr or rr in rp.parents


def _display(path: Path) -> str:
    """Workspace-relative path when possible, otherwise the absolute path."""
    try:
        return str(path.resolve().relative_to(WORKSPACE_DIR.resolve()))
    except ValueError:
        return str(path.resolve())


def _solver_files(root: Path, limit: int = 4000) -> list[Path]:
    """Non-reference files under ``root`` (the solver's outputs)."""
    reference_root = (WORKSPACE_DIR / "reference_files").resolve()
    out: list[Path] = []
    seen: set[Path] = set()
    try:
        for dirpath, dirnames, filenames in os.walk(root):
            dirnames[:] = [d for d in dirnames if d not in _WALK_SKIP_DIRS]
            for name in filenames:
                path = Path(dirpath) / name
                try:
                    if not path.is_file():  # skip sockets, fifos, broken symlinks
                        continue
                    resolved = path.resolve()
                except OSError:
                    continue
                if resolved in seen:
                    continue
                if resolved == reference_root or reference_root in resolved.parents:
                    continue
                seen.add(resolved)
                out.append(path)
                if len(out) >= limit:
                    return out
    except OSError:
        pass
    return out


def _misplaced_files(suffix: str) -> list[str]:
    """Deliverable-type files found in scratch dirs (e.g. agent wrote the .xlsx to
    /tmp). Diagnostics only, never graded; filtered to the deliverable's file type
    to stay signal-rich."""
    suffix = suffix.lower()
    found: list[str] = []
    for candidate in SCRATCH_ROOTS:
        rp = Path(candidate)
        if not rp.exists() or _within(WORKSPACE_DIR, rp):
            continue
        found.extend(
            str(p) for p in _solver_files(rp, limit=200)
            if (not suffix) or p.suffix.lower() == suffix
        )
    return found[:40]


def _resolve_deliverable(expected: Path) -> tuple[Path | None, dict[str, Any]]:
    """Resolve the agent's deliverable within the workspace, tolerating small naming
    slips, without forgiving genuinely missing work. Grading is scoped to the
    workspace only; files in scratch dirs are surfaced as diagnostics, never graded."""
    if expected.is_file():
        return expected, {}

    expected_name = expected.name.lower()
    expected_stem = expected.stem.lower()
    expected_suffix = expected.suffix.lower()
    files = _solver_files(WORKSPACE_DIR)
    same_suffix = [path for path in files if path.suffix.lower() == expected_suffix]

    def pick(matches: list[Path]) -> Path | None:
        if len(matches) == 1:
            return matches[0]
        if not matches:
            return None
        try:
            return max(matches, key=lambda p: p.stat().st_mtime)
        except OSError:
            return matches[0]

    def found(path: Path) -> tuple[Path, dict[str, Any]]:
        return path, {"deliverable_resolved_from": _display(path), "expected_deliverable": _display(expected)}

    name_match = pick([p for p in same_suffix if p.name.lower() == expected_name])
    if name_match is not None:
        return found(name_match)

    stem_match = pick([
        p for p in same_suffix
        if expected_stem in p.stem.lower() or p.stem.lower() in expected_stem
    ])
    if stem_match is not None:
        return found(stem_match)

    if len(same_suffix) == 1:
        return found(same_suffix[0])

    return None, {
        "expected_deliverable": _display(expected),
        "candidate_deliverables": [_display(path) for path in same_suffix[:20]],
        "workspace_files": [_display(path) for path in files[:80]],
        "misplaced_files_outside_workspace": _misplaced_files(expected_suffix),
    }


def _zero(status: str, **info: Any) -> dict[str, Any]:
    return {"reward": 0.0, "info": {"status": status, **info}}


def _workspace_preamble(deliverable_rel: str) -> str:
    """Factual environment contract prepended to every task prompt.

    States only *where* things live â€” the working directory, the staged input
    directory, and the absolute deliverable path â€” the same context a real analyst
    is given. It deliberately does not coach navigation or warn against fabrication:
    if an agent ignores the stated paths or invents data, that is a genuine failure
    and should score accordingly. The grader still only reads the workspace.
    """
    ws = str(WORKSPACE_DIR)
    deliverable_abs = str((WORKSPACE_DIR / deliverable_rel))
    return (
        "[Environment]\n"
        f"- Working directory: {ws}\n"
        f"- Reference files for this task are staged at: {ws}/reference_files/\n"
        f"- Save your deliverable to: {deliverable_abs}\n\n"
    )


def _stage_bundle(task_slug: str) -> None:
    """Copy tasks/<slug>/reference_files/ into the workspace. The _hidden/ key is never copied."""
    source = APP_ROOT / "tasks" / task_slug / "reference_files"
    if not source.is_dir():
        raise RuntimeError(f"no reference files staged for task {task_slug!r} at {source}")
    target = WORKSPACE_DIR / "reference_files"
    if target.exists():
        shutil.rmtree(target)
    shutil.copytree(source, target, ignore=shutil.ignore_patterns(".gitkeep"))


def _load_grader(task_slug: str, grader_source: str):
    """Load the grader from a plaintext task arg (sync-only authoring) or, if none
    is supplied, from the image file tasks/<slug>/grader.py."""
    task_dir = APP_ROOT / "tasks" / task_slug
    for path in (APP_ROOT, task_dir):
        path_text = str(path)
        if path_text not in sys.path:
            sys.path.insert(0, path_text)
    importlib.invalidate_caches()
    for helper in ("deliverable_io", "native_grading"):
        helper_path = APP_ROOT / f"{helper}.py"
        if helper not in sys.modules and helper_path.is_file():
            spec = importlib.util.spec_from_file_location(helper, helper_path)
            if spec is not None and spec.loader is not None:
                module = importlib.util.module_from_spec(spec)
                sys.modules[helper] = module
                spec.loader.exec_module(module)
    if grader_source.strip():
        module = types.ModuleType(f"gdpval_grader_arg_{abs(hash(task_slug))}")
        module.__dict__["__file__"] = str(task_dir / "grader.py")
        exec(compile(grader_source, "<grader_source>", "exec"), module.__dict__)  # noqa: S102 - task-author code
        return module
    grader_path = task_dir / "grader.py"
    spec = importlib.util.spec_from_file_location(f"gdpval_grader_{abs(hash(task_slug))}", grader_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"could not load grader at {grader_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


async def _grade(task_slug: str, deliverable: Path, grader_source: str, rubric_arg: Any) -> dict[str, Any]:
    grader = _load_grader(task_slug, grader_source)
    if isinstance(rubric_arg, dict) and rubric_arg:
        rubric = rubric_arg
    else:
        return _zero("missing_rubric_arg")
    result = grader.grade(WORKSPACE_DIR, deliverable, rubric)
    if inspect.isawaitable(result):
        result = await result
    return result if isinstance(result, dict) and "reward" in result else _zero("invalid_grader_result")


def _to_eval(result: dict[str, Any]) -> Any:
    """Return a HUD v6 evaluation while preserving grader diagnostics."""
    reward = float(result.get("reward", 0.0))
    reward = max(0.0, min(1.0, reward))
    info = result.get("info") if isinstance(result.get("info"), dict) else {}
    status = str(info.get("status", "graded"))
    subscores: list[Any] = []
    if SubScore is not None:
        for item in info.get("subscores", []):
            if isinstance(item, dict) and {"name", "value", "weight"} <= set(item):
                try:
                    subscores.append(
                        SubScore(
                            name=str(item["name"]),
                            value=max(0.0, min(1.0, float(item["value"]))),
                            weight=float(item["weight"]),
                        )
                    )
                except (TypeError, ValueError):
                    continue
    if EvaluationResult is None:
        return {"score": reward, "done": True, "content": status, "info": info}
    return EvaluationResult(
        reward=reward,
        done=True,
        content=status,
        info=info,
        isError=status.endswith("error") or status == "grader_error",
        subscores=subscores or None,
    )


async def _run(task_slug: str, prompt: str, deliverable_rel: str,
               grader_source: str, rubric_arg: Any) -> AsyncGenerator[Any, None]:
    if not task_slug:
        raise RuntimeError("task args must include task_slug")
    deliverable = _deliverable_path(deliverable_rel)
    deliverable.parent.mkdir(parents=True, exist_ok=True)
    _stage_bundle(task_slug)
    _chown_solver_workspace()

    _ = yield _workspace_preamble(deliverable_rel) + prompt

    resolved_deliverable, resolution_info = _resolve_deliverable(deliverable)
    if resolved_deliverable is None:
        yield _to_eval(_zero("missing_deliverable", **resolution_info))
        return
    try:
        result = await _grade(task_slug, resolved_deliverable, grader_source, rubric_arg)
        if resolution_info:
            result.setdefault("info", {}).update(resolution_info)
        yield _to_eval(result)
    except Exception as exc:  # pragma: no cover - surfaces author errors
        yield _to_eval(_zero("grader_error", reason=repr(exc)[:500]))



@env.template(id="gdpval_task")
async def gdpval_task(
    prompt: str = "Read the staged reference_files and produce the requested deliverable.",
    task_slug: str = "",
    deliverable: str = "deliverable/report.md",
    grader_source: str = "",
    rubric: dict[str, Any] | None = None,
) -> AsyncGenerator[Any, None]:
    # grader_source + rubric may travel as plaintext task args (sync-only authoring);
    # if omitted, the image copies under tasks/<slug>/ are used.
    async for item in _run(task_slug, prompt, deliverable, grader_source, rubric):
        yield item
